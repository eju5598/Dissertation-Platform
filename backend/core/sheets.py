import io
from openpyxl import Workbook

def build_marking_sheet_xlsx(assignment):
    wb = Workbook()
    ws = wb.active
    ws.title = "Marking Sheet"
    s = assignment.student
    p = assignment.project
    ws["A1"] = "Student Name"; ws["B1"] = s.name
    ws["A2"] = "Matriculation ID"; ws["B2"] = s.matric_id
    ws["A3"] = "Project Title"; ws["B3"] = p.title if p else ""
    ws["A4"] = "Supervisor"; ws["B4"] = assignment.supervisor.email if assignment.supervisor else ""
    ws["A5"] = "Second Marker"; ws["B5"] = assignment.second_marker.email if assignment.second_marker else ""
    bio = io.BytesIO(); wb.save(bio); bio.seek(0); return bio
